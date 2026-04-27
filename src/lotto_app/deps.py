from __future__ import annotations

from dataclasses import dataclass


@dataclass
class ExcelDeps:
    Workbook: object
    load_workbook: object
    Font: object
    font_red: object
    font_bold: object


def ensure_excel_dependencies() -> ExcelDeps:
    try:
        from openpyxl import Workbook
        from openpyxl import load_workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError(
            "missing dependency: %s. Install them with 'python -m pip install -e .'." % exc.name
        ) from exc

    return ExcelDeps(
        Workbook=Workbook,
        load_workbook=load_workbook,
        Font=Font,
        font_red=Font(color="00FF0000", bold=True),
        font_bold=Font(bold=True),
    )
